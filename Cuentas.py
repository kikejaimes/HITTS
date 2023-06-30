import pandas as pd

df = pd.read_excel('nombre_del_archivo.xlsx')

from openpyxl import load_workbook

workbook = load_workbook(filename="nombre_del_archivo.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, max_col=3):
    values = (cell.value for cell in row)
    sheet.cell(row=row[0].row, column=4).value = ''.join(values)

workbook.save(filename="nombre_del_archivo.xlsx")